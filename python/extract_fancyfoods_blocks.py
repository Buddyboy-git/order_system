def best_guess_uom(desc, code, category):
    # Only used for provisions section, and only if UOM is not found in z2/z3
    desc = str(desc or '').upper()
    code = str(code or '').upper()
    # Heuristics for common UOMs in provisions
    # If description or code contains box, bx, cs, case, pk, pkg, default to CS or BX
    # Should never be called, but return 'LB' as a safe fallback
    return 'LB'
import openpyxl
import csv
import sys

EXCEL_PATH = r'd:/xampp/htdocs/order_system/data/excel_price_sheets/fancyfoods/FancyFoods Price Sheet.xlsx'
CSV_PATH = r'd:/xampp/htdocs/order_system/data/excel_price_sheets/fancyfoods/fancyfoods_extracted.csv'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

output_header = ['Category', 'Subcategory', 'Item Code', 'Description', 'Price', 'UOM']
output_rows = []

max_row = ws.max_row
max_col = ws.max_column
row = 1
while row <= max_row:
    # Find header row for a block (must contain CODE, DESCRIPTION, Z1)
    header_cells = [str(ws.cell(row=row, column=col).value).strip().upper() if ws.cell(row=row, column=col).value else '' for col in range(1, max_col+1)]
    if 'CODE' in header_cells and 'DESCRIPTION' in header_cells and 'Z1' in header_cells:
        print(f"[DEBUG] Header detected at row {row}", file=sys.stderr)
        # Find category label exactly one row above header row
        cat_row = row - 1
        category = 'UNCATEGORIZED'
        if cat_row >= 1:
            cat_values = [str(ws.cell(row=cat_row, column=col).value).strip() if ws.cell(row=cat_row, column=col).value else '' for col in range(1, max_col+1)]
            for val in cat_values:
                val_lower = val.lower()
                if 'price sheet' in val_lower:
                    idx = val_lower.find('price sheet')
                    category_candidate = val[:idx].strip()
                    if category_candidate:
                        category = category_candidate
                    break
        # Find left and right block column indices
        left = {}
        right = {}
        # Dynamically detect all relevant columns for left and right blocks
        for col in range(1, max_col+1):
            val = str(ws.cell(row=row, column=col).value).strip().upper() if ws.cell(row=row, column=col).value else ''
            # Left block
            if not right and (not left.get('code') or not left.get('desc') or not left.get('z1') or not left.get('z2') or not left.get('z3')):
                if val == 'CODE' and not left.get('code'):
                    left['code'] = col
                elif val == 'DESCRIPTION' and not left.get('desc'):
                    left['desc'] = col
                elif val == 'Z1' and not left.get('z1'):
                    left['z1'] = col
                elif val == 'Z2' and not left.get('z2'):
                    left['z2'] = col
                elif val == 'Z3' and not left.get('z3'):
                    left['z3'] = col
            # Right block
            else:
                if val == 'CODE' and not right.get('code'):
                    right['code'] = col
                elif val == 'DESCRIPTION' and not right.get('desc'):
                    right['desc'] = col
                elif val == 'Z1' and not right.get('z1'):
                    right['z1'] = col
                elif val == 'Z2' and not right.get('z2'):
                    right['z2'] = col
                elif val == 'Z3' and not right.get('z3'):
                    right['z3'] = col
        # Extract left block with subcategory detection
        r = row + 1
        empty_count = 0
        subcategory = ''
        while r <= max_row:
            header_check = [str(ws.cell(row=r, column=col).value).strip().upper() if ws.cell(row=r, column=col).value else '' for col in range(1, max_col+1)]
            if 'CODE' in header_check and 'DESCRIPTION' in header_check and 'Z1' in header_check:
                print(f"[DEBUG] New header detected at row {r} (ending left block for {category})", file=sys.stderr)
                break
            code = ws.cell(row=r, column=left['code']).value
            desc_cell = ws.cell(row=r, column=left['desc'])
            desc = desc_cell.value
            price = None
            uom = None
            z1_val = ws.cell(row=r, column=left['z1']).value if 'z1' in left else None
            z2_val = ws.cell(row=r, column=left['z2']).value if 'z2' in left else None
            z3_val = ws.cell(row=r, column=left['z3']).value if 'z3' in left else None
            # Price selection logic (literal)
            if z1_val is not None and str(z1_val).strip() != '':
                price = z1_val
                # UOM is z2 if present and is a non-numeric, non-empty string
                if 'z2' in left:
                    z2_uom_candidate = ws.cell(row=r, column=left['z2']).value
                    if isinstance(z2_uom_candidate, str) and z2_uom_candidate.strip():
                        try:
                            float(z2_uom_candidate)
                        except ValueError:
                            uom = z2_uom_candidate.strip().upper()
            elif z2_val is not None and str(z2_val).strip() != '':
                price = z2_val
                # UOM is z3 if present and is a non-numeric, non-empty string
                if 'z3' in left:
                    z3_uom_candidate = ws.cell(row=r, column=left['z3']).value
                    if isinstance(z3_uom_candidate, str) and z3_uom_candidate.strip():
                        try:
                            float(z3_uom_candidate)
                        except ValueError:
                            uom = z3_uom_candidate.strip().upper()
            elif z3_val is not None and str(z3_val).strip() != '':
                price = z3_val
            # Detect yellow fill for subcategory (openpyxl uses RGB, yellow is usually 'FFFF00')
            # Detect yellow fill for subcategory (openpyxl uses RGB, yellow is usually 'FFFF00')
            fill = desc_cell.fill
            is_yellow = False
            if fill and hasattr(fill, 'fgColor') and fill.fgColor.type == 'rgb':
                is_yellow = fill.fgColor.rgb in ['FFFF00', 'FFFFFF00']
            # Only check if code is empty for subcategory detection
            if is_yellow and desc and not code:
                subcategory = str(desc).strip()
                print(f"[DEBUG] Subcategory detected: {subcategory} at row {r}", file=sys.stderr)
            elif not code and not desc:
                empty_count += 1
                if empty_count >= 5:
                    print(f"[DEBUG] 5 consecutive empty rows at row {r} (ending left block for {category})", file=sys.stderr)
                    break
            else:
                empty_count = 0
                if code and desc:
                    print(f"[DEBUG] Extracted LEFT: {category}, {subcategory}, {str(code).strip()}, {str(desc).strip()}, {str(price).strip() if price else ''}", file=sys.stderr)
                    # UOM logic: use non-numeric from z2/z3 if set, else default to LB
                    if not uom:
                        uom = 'LB'
                    output_rows.append([category, subcategory, str(code).strip(), str(desc).strip(), str(price).strip() if price else '', uom])
            r += 1
        # Extract right block with subcategory detection
        r = row + 1
        empty_count = 0
        subcategory = ''
        while r <= max_row:
            header_check = [str(ws.cell(row=r, column=col).value).strip().upper() if ws.cell(row=r, column=col).value else '' for col in range(1, max_col+1)]
            if 'CODE' in header_check and 'DESCRIPTION' in header_check and 'Z1' in header_check:
                print(f"[DEBUG] New header detected at row {r} (ending right block for {category})", file=sys.stderr)
                break
            code = ws.cell(row=r, column=right['code']).value if 'code' in right else None
            desc_cell = ws.cell(row=r, column=right['desc']) if 'desc' in right else None
            desc = desc_cell.value if desc_cell else None
            price = None
            uom = None
            z1_val = ws.cell(row=r, column=right['z1']).value if 'z1' in right else None
            z2_val = ws.cell(row=r, column=right['z2']).value if 'z2' in right else None
            z3_val = ws.cell(row=r, column=right['z3']).value if 'z3' in right else None
            # Price selection logic (literal)
            if z1_val is not None and str(z1_val).strip() != '':
                price = z1_val
                # UOM is z2 if present and is a non-numeric, non-empty string
                if 'z2' in right:
                    z2_uom_candidate = ws.cell(row=r, column=right['z2']).value
                    if isinstance(z2_uom_candidate, str) and z2_uom_candidate.strip():
                        try:
                            float(z2_uom_candidate)
                        except ValueError:
                            uom = z2_uom_candidate.strip().upper()
            elif z2_val is not None and str(z2_val).strip() != '':
                price = z2_val
                # UOM is z3 if present and is a non-numeric, non-empty string
                if 'z3' in right:
                    z3_uom_candidate = ws.cell(row=r, column=right['z3']).value
                    if isinstance(z3_uom_candidate, str) and z3_uom_candidate.strip():
                        try:
                            float(z3_uom_candidate)
                        except ValueError:
                            uom = z3_uom_candidate.strip().upper()
            elif z3_val is not None and str(z3_val).strip() != '':
                price = z3_val
            fill = desc_cell.fill if desc_cell else None
            is_yellow = False
            if fill and hasattr(fill, 'fgColor') and fill.fgColor.type == 'rgb':
                is_yellow = fill.fgColor.rgb in ['FFFF00', 'FFFFFF00']
            if is_yellow and desc and not code:
                subcategory = str(desc).strip()
                print(f"[DEBUG] Subcategory detected: {subcategory} at row {r}", file=sys.stderr)
            elif not code and not desc:
                empty_count += 1
                if empty_count >= 5:
                    print(f"[DEBUG] 5 consecutive empty rows at row {r} (ending right block for {category})", file=sys.stderr)
                    break
            else:
                empty_count = 0
                if code and desc:
                    print(f"[DEBUG] Extracted RIGHT: {category}, {subcategory}, {str(code).strip()}, {str(desc).strip()}, {str(price).strip() if price else ''}", file=sys.stderr)
                    # UOM logic: use non-numeric from z2/z3 if set, else default to LB
                    if not uom:
                        uom = 'LB'
                    output_rows.append([category, subcategory, str(code).strip(), str(desc).strip(), str(price).strip() if price else '', uom])
            r += 1
        row = r
    else:
        row += 1

with open(CSV_PATH, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(output_header)
    writer.writerows(output_rows)

print(f'Extraction complete. Output written to {CSV_PATH}')
